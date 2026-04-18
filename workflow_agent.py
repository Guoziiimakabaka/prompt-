import argparse
import hashlib
import json
import os
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from dotenv import load_dotenv
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from app import extract_page_content


INPUT_HEADER_ALIASES: dict[str, list[str]] = {
    "uid": ["uid"],
    "prompt": ["prompt"],
    "repo": ["repo"],
    "scores": ["评分", "打分", "score"],
    "remarks": ["产物备注（参考）", "产物备注(参考)", "备注", "remark"],
}

OUTPUT_COLUMNS: dict[str, str] = {
    "debug_prompt": "debug prompt",
    "debug_repo": "debug 改写repo",
    "debug_unable_note": "无法写debug改写备注",
    "new_prompt": "新增需求 prompt",
    "new_repo": "新增需求改写 repo",
    "new_unable_note": "无法新增需求备注",
}

DEBUG_UNABLE_TEXT = "7 个产物都很优秀，无法写 debug"


@dataclass
class ParsedInput:
    uid: str
    base_prompt: str
    repo_urls: list[str]
    scores: list[int]
    remarks: dict[int, str]


@dataclass
class RepoResult:
    repo_id: str
    url: str
    score: int
    remark: str
    ok: bool
    title: str
    text_preview: str
    text_length: int
    error: str


@dataclass
class RuntimeCheck:
    repo_id: str
    checked: bool
    passed: bool
    issues: list[str]
    behavior_checked: bool
    behavior_passed: bool
    behavior_issues: list[str]
    behavior_evidence: list[dict[str, Any]]
    remark_probe_checked: bool
    remark_related_issues: list[str]


@dataclass
class AnalysisBundle:
    repo_issues: dict[str, list[str]]
    global_issues: list[str]
    runtime_checks: dict[str, RuntimeCheck]


class AgentDecision(BaseModel):
    debug_prompt: str = Field(...)
    debug_rewrite_repo: str = Field(...)
    new_requirement_prompt: str = Field(...)
    new_requirement_rewrite_repo: str = Field(...)
    unable_to_add_requirement_note: str = Field(...)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read rows from Excel, run fetch + behavior-driven runtime validation, "
            "generate prompts, and write results back to Excel."
        )
    )
    parser.add_argument("--input_xlsx", default="input.xlsx", help="Input Excel path.")
    parser.add_argument(
        "--output_xlsx",
        default="",
        help="Output Excel path. Empty means overwrite input file.",
    )
    parser.add_argument(
        "--sheet_name",
        default="",
        help="Sheet name to process. Empty means first sheet.",
    )
    parser.add_argument(
        "--start_row",
        type=int,
        default=2,
        help="Start row index (1-based).",
    )
    parser.add_argument(
        "--max_rows",
        type=int,
        default=0,
        help="Max rows to process. 0 means no limit.",
    )
    parser.add_argument(
        "--uid",
        default="",
        help="Only process rows whose uid exactly matches.",
    )
    parser.add_argument(
        "--fetch_json",
        default="repo_fetch_results.json",
        help="Output JSON path for fetch/runtime details.",
    )
    return parser.parse_args()


def _to_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize_header(text: str) -> str:
    cleaned = _to_text(text)
    cleaned = cleaned.replace("（", "(").replace("）", ")")
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned.lower()


def _require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise ValueError(f"{name} must be set in .env")
    return value


def _raw_header_map(ws: Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = _to_text(ws.cell(row=1, column=col).value)
        if header:
            mapping[header] = col
    return mapping


def _resolve_input_columns(raw_map: dict[str, int]) -> dict[str, int]:
    normalized_to_col = {_normalize_header(k): v for k, v in raw_map.items()}
    resolved: dict[str, int] = {}
    for key, aliases in INPUT_HEADER_ALIASES.items():
        for alias in aliases:
            norm = _normalize_header(alias)
            if norm in normalized_to_col:
                resolved[key] = normalized_to_col[norm]
                break
    missing = [k for k in INPUT_HEADER_ALIASES if k not in resolved]
    if missing:
        raise ValueError(f"Missing required input columns: {missing}")
    return resolved


def _ensure_output_columns(ws: Worksheet, raw_map: dict[str, int]) -> dict[str, int]:
    result: dict[str, int] = {}
    next_col = ws.max_column + 1
    normalized_to_col = {_normalize_header(k): v for k, v in raw_map.items()}
    for key, header in OUTPUT_COLUMNS.items():
        norm = _normalize_header(header)
        if norm in normalized_to_col:
            result[key] = normalized_to_col[norm]
        else:
            ws.cell(row=1, column=next_col).value = header
            result[key] = next_col
            next_col += 1
    return result


def _parse_json_list(raw: str, field_name: str) -> list[Any]:
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{field_name} must be valid JSON list: {raw}") from exc
    if not isinstance(parsed, list):
        raise ValueError(f"{field_name} must be a JSON list")
    return parsed


def _parse_urls(raw: str) -> list[str]:
    values = _parse_json_list(raw, "repo")
    if len(values) != 7:
        raise ValueError("repo URL list must contain exactly 7 URLs")
    urls: list[str] = []
    for value in values:
        text = _to_text(value)
        if not text.startswith("http"):
            raise ValueError(f"Invalid URL in repo list: {text}")
        urls.append(text)
    return urls


def _parse_scores(raw: str) -> list[int]:
    values = _parse_json_list(raw, "评分")
    if len(values) != 7:
        raise ValueError("评分 list must contain exactly 7 items")
    scores: list[int] = []
    for value in values:
        score = int(_to_text(value))
        if score not in {0, 1, 2}:
            raise ValueError(f"评分 must contain only 0/1/2, got: {score}")
        scores.append(score)
    return scores


def _parse_remarks(raw: str) -> dict[int, str]:
    if not raw:
        return {}
    normalized = raw.replace("；", "\n").replace(";", "\n")
    remarks: dict[int, str] = {}
    for line in normalized.splitlines():
        text = line.strip()
        if not text:
            continue
        match = re.match(r"repo\s*([1-7])\s*[:：]\s*(.*)$", text, re.IGNORECASE)
        if match:
            index = int(match.group(1))
            note = match.group(2).strip()
            if note:
                remarks[index] = note
    return remarks


def parse_row_input(
    ws: Worksheet,
    row_idx: int,
    columns: dict[str, int],
) -> ParsedInput | None:
    uid = _to_text(ws.cell(row=row_idx, column=columns["uid"]).value)
    prompt = _to_text(ws.cell(row=row_idx, column=columns["prompt"]).value)
    repo_raw = _to_text(ws.cell(row=row_idx, column=columns["repo"]).value)
    scores_raw = _to_text(ws.cell(row=row_idx, column=columns["scores"]).value)
    remarks_raw = _to_text(ws.cell(row=row_idx, column=columns["remarks"]).value)

    if not uid and not prompt and not repo_raw and not scores_raw:
        return None
    if not uid:
        raise ValueError(f"Row {row_idx}: uid is empty")
    if not prompt:
        raise ValueError(f"Row {row_idx}: prompt is empty")
    if not repo_raw:
        raise ValueError(f"Row {row_idx}: repo is empty")
    if not scores_raw:
        raise ValueError(f"Row {row_idx}: 评分 is empty")

    return ParsedInput(
        uid=uid,
        base_prompt=prompt,
        repo_urls=_parse_urls(repo_raw),
        scores=_parse_scores(scores_raw),
        remarks=_parse_remarks(remarks_raw),
    )


def _repo_by_id(repo_id: str, results: list[RepoResult]) -> RepoResult | None:
    for result in results:
        if result.repo_id == repo_id:
            return result
    return None


def _is_meaningful_remark(text: str) -> bool:
    cleaned = text.strip().lower()
    return cleaned not in {"", "无", "none", "n/a", "na"}


def fetch_repos(data: ParsedInput) -> list[RepoResult]:
    results: list[RepoResult] = []
    for i, url in enumerate(data.repo_urls, start=1):
        repo_id = f"repo{i}"
        score = data.scores[i - 1]
        remark = data.remarks.get(i, "")
        try:
            extracted = extract_page_content(url)
            text = _to_text(extracted.get("text", ""))
            results.append(
                RepoResult(
                    repo_id=repo_id,
                    url=url,
                    score=score,
                    remark=remark,
                    ok=True,
                    title=_to_text(extracted.get("title", "")),
                    text_preview=text[:500].replace("\n", " "),
                    text_length=int(extracted.get("text_length", 0)),
                    error="",
                )
            )
        except Exception as exc:
            results.append(
                RepoResult(
                    repo_id=repo_id,
                    url=url,
                    score=score,
                    remark=remark,
                    ok=False,
                    title="",
                    text_preview="",
                    text_length=0,
                    error=str(exc),
                )
            )
    return results


def _install_mutation_observer(page: Any) -> None:
    page.add_init_script(
        """
        () => {
          window.__qaMutationCount = 0;
          const observer = new MutationObserver((records) => {
            window.__qaMutationCount += records.length;
          });
          observer.observe(document, {
            subtree: true,
            childList: true,
            attributes: true,
            characterData: true
          });
        }
        """
    )


def _dom_snapshot(page: Any) -> tuple[str, int, int]:
    payload = page.evaluate(
        """
        () => {
          const body = document.body;
          const text = body ? (body.innerText || '').slice(0, 8000) : '';
          const html = body ? (body.innerHTML || '').slice(0, 16000) : '';
          const mutation = Number(window.__qaMutationCount || 0);
          const interactive = document.querySelectorAll(
            "button, [role='button'], input, select, textarea, a, [onclick]"
          ).length;
          return { text, html, mutation, interactive };
        }
        """
    )
    seed = f"{payload['text']}||{payload['html']}"
    digest = hashlib.sha1(seed.encode("utf-8", errors="ignore")).hexdigest()
    return digest, int(payload["mutation"]), int(payload["interactive"])


def _element_meta(element: Any) -> dict[str, str]:
    return element.evaluate(
        """
        (el) => ({
          tag: (el.tagName || '').toLowerCase(),
          id: el.id || '',
          cls: (el.className || '').toString().slice(0, 120),
          text: ((el.innerText || el.value || '').trim()).slice(0, 40),
          href: el.getAttribute ? (el.getAttribute('href') || '') : ''
        })
        """
    )


def _same_origin(url: str, href: str) -> bool:
    if not href:
        return True
    if href.startswith("#") or href.startswith("javascript:"):
        return True
    if href.startswith("/"):
        return True
    if not href.startswith("http"):
        return True
    return urlparse(url).netloc == urlparse(href).netloc


def _collect_click_candidates(page: Any, current_url: str) -> list[Any]:
    selector = (
        "button, [role='button'], input[type='button'], input[type='submit'], "
        "a, [onclick], [data-testid*='cell'], [class*='cell'], [class*='tile'], "
        "[class*='square'], [role='gridcell']"
    )
    elements = page.query_selector_all(selector)
    candidates: list[Any] = []
    for element in elements:
        try:
            if not element.is_visible():
                continue
            if not element.is_enabled():
                continue
            meta = _element_meta(element)
            if meta["tag"] == "a" and not _same_origin(current_url, meta["href"]):
                continue
            if not meta["text"] and not meta["id"] and not meta["cls"]:
                continue
            candidates.append(element)
            if len(candidates) >= 24:
                break
        except Exception:
            continue
    return candidates


def _run_generic_interaction_probe(page: Any, current_url: str) -> tuple[list[str], dict[str, Any]]:
    issues: list[str] = []
    evidence: dict[str, Any] = {
        "probe": "generic_interaction",
        "attempts": 0,
        "state_changes": 0,
        "clicked": [],
    }
    candidates = _collect_click_candidates(page, current_url)
    if not candidates:
        issues.append("未发现可交互控件，页面缺少可验证行为。")
        return issues, evidence

    baseline_sig, baseline_mutation, _ = _dom_snapshot(page)
    for element in candidates[:8]:
        evidence["attempts"] += 1
        try:
            meta = _element_meta(element)
            element.scroll_into_view_if_needed(timeout=1200)
            element.click(timeout=1800)
            page.wait_for_timeout(250)
            now_sig, now_mutation, _ = _dom_snapshot(page)
            changed = now_sig != baseline_sig or now_mutation > baseline_mutation
            evidence["clicked"].append(
                {
                    "tag": meta["tag"],
                    "text": meta["text"],
                    "id": meta["id"],
                    "changed": changed,
                }
            )
            if changed:
                evidence["state_changes"] += 1
                baseline_sig = now_sig
                baseline_mutation = now_mutation
        except Exception as exc:
            evidence["clicked"].append({"error": str(exc)[:160], "changed": False})

    if evidence["state_changes"] == 0:
        issues.append("执行多次真实点击后页面状态无变化，疑似交互逻辑未生效。")
    return issues, evidence


def _run_form_probe(page: Any) -> tuple[list[str], dict[str, Any]]:
    issues: list[str] = []
    evidence: dict[str, Any] = {"probe": "form_input", "input_changed": False}
    try:
        target = page.query_selector(
            "input[type='text'], input:not([type]), textarea, input[type='search']"
        )
        if target and target.is_visible() and target.is_enabled():
            target.fill("qa-behavior-probe")
            page.wait_for_timeout(120)
            value = target.input_value()
            evidence["value"] = value
            evidence["input_changed"] = "qa-behavior-probe" in value
            if not evidence["input_changed"]:
                issues.append("输入控件填充后值未变化，疑似表单输入链路异常。")
        else:
            evidence["skipped"] = True
    except Exception as exc:
        issues.append(f"表单探测执行异常：{exc}")
    return issues, evidence


def _is_board_like_prompt(base_prompt: str) -> bool:
    return bool(
        re.search(
            r"boop|kitten|kittens|cats|board|6x6|棋盘|落子|回合|胜负",
            base_prompt,
            re.IGNORECASE,
        )
    )


def _collect_board_cells(page: Any) -> list[Any]:
    selector = (
        "[data-cell], [data-testid*='cell'], [class*='cell'], [class*='tile'], "
        "[class*='square'], [role='gridcell'], td"
    )
    elements = page.query_selector_all(selector)
    cells: list[Any] = []
    for element in elements:
        try:
            if not element.is_visible():
                continue
            if not element.is_enabled():
                continue
            box = element.bounding_box()
            if not box:
                continue
            area = box["width"] * box["height"]
            if area < 80 or area > 70000:
                continue
            cells.append(element)
            if len(cells) >= 40:
                break
        except Exception:
            continue
    return cells


def _run_board_probe(page: Any, base_prompt: str) -> tuple[list[str], dict[str, Any]]:
    issues: list[str] = []
    evidence: dict[str, Any] = {
        "probe": "board_interaction",
        "attempts": 0,
        "state_changes": 0,
        "cell_count": 0,
    }
    if not _is_board_like_prompt(base_prompt):
        evidence["skipped"] = True
        return issues, evidence

    cells = _collect_board_cells(page)
    evidence["cell_count"] = len(cells)
    if len(cells) < 9:
        issues.append("页面疑似棋盘类场景，但未识别到足够棋盘单元。")
        return issues, evidence

    baseline_sig, baseline_mutation, _ = _dom_snapshot(page)
    for cell in cells[:12]:
        evidence["attempts"] += 1
        try:
            cell.scroll_into_view_if_needed(timeout=1000)
            cell.click(timeout=1500)
            page.wait_for_timeout(220)
            now_sig, now_mutation, _ = _dom_snapshot(page)
            changed = now_sig != baseline_sig or now_mutation > baseline_mutation
            if changed:
                evidence["state_changes"] += 1
                baseline_sig = now_sig
                baseline_mutation = now_mutation
                if evidence["state_changes"] >= 2:
                    break
        except Exception:
            continue

    if evidence["state_changes"] == 0:
        issues.append("棋盘候选单元多次点击后未观察到状态变化，疑似落子逻辑未生效。")
    return issues, evidence


def _classify_remark_tokens(remark: str) -> dict[str, bool]:
    text = remark.lower()
    return {
        "title_or_template": bool(
            re.search(r"标题|react app|模板|template|title|页签|tab", text)
        ),
        "icon_or_style": bool(
            re.search(r"图标|icon|样式|布局|丑|美观|未分开|混淆", text)
        ),
        "interaction_or_mechanic": bool(
            re.search(r"机制|规则|交互|点击|卡死|死局|无法|回收|胜利|连线|落子", text)
        ),
        "progress_or_step": bool(
            re.search(r"进度|步骤|stage|flow|流程", text)
        ),
    }


def _probe_remark_specific_issues(
    remark: str,
    title: str,
    page_errors: list[str],
    console_errors: list[str],
    behavior_issues: list[str],
    behavior_evidence: list[dict[str, Any]],
) -> tuple[list[str], dict[str, Any]]:
    if not _is_meaningful_remark(remark):
        return [], {"skipped": True}

    tags = _classify_remark_tokens(remark)
    issues: list[str] = []
    evidence: dict[str, Any] = {
        "probe": "remark_directed",
        "remark": remark[:200],
        "tags": tags,
    }

    if tags["title_or_template"] and title.strip().strip('"') in {"", "React App"}:
        issues.append("备注涉及标题/模板问题，实测页面标题仍为默认模板或为空。")

    if tags["progress_or_step"]:
        generic_probe = next(
            (item for item in behavior_evidence if item.get("probe") == "generic_interaction"),
            {},
        )
        attempts = int(generic_probe.get("attempts", 0))
        changes = int(generic_probe.get("state_changes", 0))
        evidence["progress_probe"] = {"attempts": attempts, "state_changes": changes}
        if attempts > 0 and changes == 0:
            issues.append("备注涉及流程/进度问题，实测多次交互后流程状态无变化。")

    if tags["interaction_or_mechanic"]:
        if behavior_issues:
            issues.append("备注涉及机制/交互问题，实测行为探测出现可复现异常。")
        if page_errors or console_errors:
            issues.append("备注涉及机制/交互问题，实测存在脚本或控制台错误。")

    if tags["icon_or_style"]:
        # 对视觉类问题仅在存在可观测异常时给出可确认结论，不做主观评价。
        if page_errors or console_errors:
            issues.append("备注涉及视觉表现问题，实测存在前端错误，可能导致界面呈现异常。")

    deduped: list[str] = []
    for item in issues:
        if item not in deduped:
            deduped.append(item)
    return deduped, evidence


def run_runtime_check(repo_id: str, url: str, base_prompt: str, remark: str) -> RuntimeCheck:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        missing = f"运行检查不可用：未安装 Playwright。{exc}"
        return RuntimeCheck(
            repo_id=repo_id,
            checked=False,
            passed=False,
            issues=[missing],
            behavior_checked=False,
            behavior_passed=False,
            behavior_issues=[missing],
            behavior_evidence=[],
            remark_probe_checked=False,
            remark_related_issues=[],
        )

    page_errors: list[str] = []
    console_errors: list[str] = []
    failed_requests: list[str] = []
    health_issues: list[str] = []
    behavior_issues: list[str] = []
    behavior_evidence: list[dict[str, Any]] = []
    remark_related_issues: list[str] = []
    remark_probe_checked = False

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(ignore_https_errors=True)
            page = context.new_page()
            _install_mutation_observer(page)

            page.on("pageerror", lambda exc: page_errors.append(str(exc)))
            page.on(
                "console",
                lambda msg: console_errors.append(msg.text)
                if msg.type == "error"
                else None,
            )
            page.on("requestfailed", lambda req: failed_requests.append(req.url))

            page.goto(url, wait_until="networkidle", timeout=35000)
            page.wait_for_timeout(350)

            root_rendered = page.evaluate(
                """
                () => {
                  const root = document.querySelector('#root');
                  if (!root) return false;
                  return (
                    root.childElementCount > 0 ||
                    (root.textContent || '').trim().length > 0
                  );
                }
                """
            )
            body_text = page.inner_text("body")
            body_len = len((body_text or "").strip())

            if not root_rendered and body_len < 30:
                health_issues.append("页面运行后接近空白，疑似渲染失败。")
            if page_errors:
                health_issues.append(f"捕获到页面脚本异常 {len(page_errors)} 条。")
            if console_errors:
                health_issues.append(f"捕获到控制台错误 {len(console_errors)} 条。")
            if len(failed_requests) >= 3:
                health_issues.append(f"捕获到网络请求失败 {len(failed_requests)} 条。")

            generic_issues, generic_evidence = _run_generic_interaction_probe(page, url)
            behavior_issues.extend(generic_issues)
            behavior_evidence.append(generic_evidence)

            form_issues, form_evidence = _run_form_probe(page)
            behavior_issues.extend(form_issues)
            behavior_evidence.append(form_evidence)

            board_issues, board_evidence = _run_board_probe(page, base_prompt)
            behavior_issues.extend(board_issues)
            behavior_evidence.append(board_evidence)

            directed_issues, directed_evidence = _probe_remark_specific_issues(
                remark=remark,
                title=page.title(),
                page_errors=page_errors,
                console_errors=console_errors,
                behavior_issues=behavior_issues,
                behavior_evidence=behavior_evidence,
            )
            remark_related_issues.extend(directed_issues)
            behavior_evidence.append(directed_evidence)
            remark_probe_checked = not directed_evidence.get("skipped", False)

            context.close()
            browser.close()
    except Exception as exc:
        err = f"运行检查执行失败：{exc}"
        return RuntimeCheck(
            repo_id=repo_id,
            checked=False,
            passed=False,
            issues=[err],
            behavior_checked=False,
            behavior_passed=False,
            behavior_issues=[err],
            behavior_evidence=behavior_evidence,
            remark_probe_checked=remark_probe_checked,
            remark_related_issues=remark_related_issues,
        )

    all_issues = health_issues + behavior_issues + remark_related_issues
    return RuntimeCheck(
        repo_id=repo_id,
        checked=True,
        passed=not all_issues,
        issues=all_issues,
        behavior_checked=True,
        behavior_passed=not behavior_issues,
        behavior_issues=behavior_issues,
        behavior_evidence=behavior_evidence,
        remark_probe_checked=remark_probe_checked,
        remark_related_issues=remark_related_issues,
    )


def build_analysis(results: list[RepoResult], base_prompt: str) -> AnalysisBundle:
    repo_issues: dict[str, list[str]] = {}
    global_issues: list[str] = []
    runtime_checks: dict[str, RuntimeCheck] = {}

    failed_fetch: list[str] = []
    default_title: list[str] = []
    empty_text: list[str] = []
    behavior_failed: list[str] = []

    for result in results:
        issues: list[str] = []
        if not result.ok:
            failed_fetch.append(result.repo_id)
            issues.append(f"页面访问或解析失败：{result.error or '未知错误'}")
        if result.title.strip().strip('"') in {"", "React App"}:
            default_title.append(result.repo_id)
            issues.append("页面标题疑似默认模板。")
        if result.text_length == 0:
            empty_text.append(result.repo_id)

        runtime = run_runtime_check(result.repo_id, result.url, base_prompt, result.remark)
        runtime_checks[result.repo_id] = runtime
        if not runtime.passed:
            issues.extend(runtime.issues)
        if runtime.checked and not runtime.behavior_passed:
            behavior_failed.append(result.repo_id)
        # “正文提取为空”通常是抓取层信号，不应单独主导 debug。
        # 仅在运行/行为也异常时，才将其升级为可行动问题。
        if result.text_length == 0 and (not runtime.passed or not runtime.behavior_passed):
            issues.append("正文提取为空，且页面运行行为异常。")

        if issues:
            repo_issues[result.repo_id] = issues

    if failed_fetch:
        global_issues.append(f"抓取失败页面：{', '.join(failed_fetch)}。")
    if default_title:
        global_issues.append(f"默认标题页面：{', '.join(default_title)}。")
    if empty_text:
        global_issues.append(f"正文为空页面：{', '.join(empty_text)}。")
    if behavior_failed:
        global_issues.append(f"行为验证未通过页面：{', '.join(behavior_failed)}。")

    return AnalysisBundle(
        repo_issues=repo_issues,
        global_issues=global_issues,
        runtime_checks=runtime_checks,
    )


def _has_repo_marker(text: str) -> bool:
    return bool(re.search(r"repo\s*\d+", text, re.IGNORECASE))


def _has_english(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]", text))


def _has_colloquial(text: str) -> bool:
    markers = ["帮我", "请你", "搞一个", "弄一个", "来个", "给我", "整一个"]
    return any(marker in text for marker in markers)


def rewrite_prompt_to_formal_chinese(
    llm: ChatOpenAI,
    prompt_text: str,
    purpose: str,
) -> str:
    system_prompt = (
        "你是提示词改写助手。"
        "请把输入改写为正式、具体、可执行的中文提示词。"
        "不得出现英文、repo编号、网址、口语化表达。"
        "只返回改写后的文本。"
    )
    human_prompt = f"用途：{purpose}\n原文：{prompt_text}"
    result = llm.invoke(
        [SystemMessage(content=system_prompt), HumanMessage(content=human_prompt)]
    )
    return _to_text(result.content)


def _sort_repos(items: list[RepoResult]) -> list[RepoResult]:
    return sorted(items, key=lambda x: (x.score, x.repo_id))


def _has_runtime_bug(repo_id: str, analysis: AnalysisBundle) -> bool:
    runtime = analysis.runtime_checks.get(repo_id)
    return bool(runtime and runtime.checked and (not runtime.passed))


def _verified_bug_lines(repo: RepoResult, analysis: AnalysisBundle) -> list[str]:
    runtime = analysis.runtime_checks.get(repo.repo_id)
    lines: list[str] = []
    if runtime:
        for item in runtime.remark_related_issues:
            text = _to_text(item)
            if text and text not in lines:
                lines.append(text)
        for item in runtime.behavior_issues:
            text = _to_text(item)
            if text and text not in lines:
                lines.append(text)
        for item in runtime.issues:
            text = _to_text(item)
            if text and text not in lines:
                lines.append(text)
    for item in analysis.repo_issues.get(repo.repo_id, []):
        text = _to_text(item)
        if text and text not in lines:
            lines.append(text)
    return lines


def _fallback_debug_prompt(
    results: list[RepoResult],
    debug_repo: str,
    analysis: AnalysisBundle,
) -> str:
    selected = _repo_by_id(debug_repo, results)
    if selected is None:
        return DEBUG_UNABLE_TEXT

    verified = _verified_bug_lines(selected, analysis)
    if not verified:
        return DEBUG_UNABLE_TEXT

    top_lines = verified[:3]
    line_text = "；".join(top_lines)
    return (
        "请基于以下已验证缺陷进行修复："
        f"{line_text}。"
        "请逐条给出对应的最小复现步骤、根因分析和修复方案，"
        "并提供同路径回归验证结果。"
    )


def _fallback_new_requirement_prompt() -> str:
    return (
        "在不改变现有核心规则的前提下，新增“操作历史回放”功能。"
        "功能要求：记录每一步操作的落点、受影响棋子与结果状态，支持逐步回放。"
        "验收标准：随机选取三局对局进行校验，回放内容与实际执行结果逐步一致，"
        "且新增功能不影响现有回合切换和胜负判定。"
    )


def _pick_debug_repo(results: list[RepoResult], analysis: AnalysisBundle) -> str:
    ordered = _sort_repos(results)
    bug_repos = [
        r
        for r in ordered
        if _has_runtime_bug(r.repo_id, analysis) and _verified_bug_lines(r, analysis)
    ]
    bug_with_remark = [r for r in bug_repos if _is_meaningful_remark(r.remark)]
    if bug_with_remark:
        return bug_with_remark[0].repo_id

    bug_score01 = [r for r in bug_repos if r.score in {0, 1}]
    if bug_score01:
        return bug_score01[0].repo_id

    remark_repos = [
        r for r in ordered if _is_meaningful_remark(r.remark) and _verified_bug_lines(r, analysis)
    ]
    if remark_repos:
        return remark_repos[0].repo_id

    score0 = [r for r in ordered if r.score == 0 and _verified_bug_lines(r, analysis)]
    if score0:
        return score0[0].repo_id

    score1 = [r for r in ordered if r.score == 1 and _verified_bug_lines(r, analysis)]
    if score1:
        return score1[0].repo_id
    return ""


def _eligible_score2_repo_ids(results: list[RepoResult], analysis: AnalysisBundle) -> list[str]:
    eligible: list[str] = []
    for result in _sort_repos([r for r in results if r.score == 2]):
        runtime = analysis.runtime_checks.get(result.repo_id)
        if (
            runtime
            and runtime.checked
            and runtime.passed
            and runtime.behavior_checked
            and runtime.behavior_passed
        ):
            eligible.append(result.repo_id)
    return eligible


def _summarize_score2_failures(results: list[RepoResult], analysis: AnalysisBundle) -> str:
    score2_repos = _sort_repos([r for r in results if r.score == 2])
    if not score2_repos:
        return "没有可用的 2 分 repo，无法新增需求。"

    details: list[str] = []
    low_signal = {
        "正文提取为空，可能依赖前端渲染。",
    }

    for item in score2_repos:
        runtime = analysis.runtime_checks.get(item.repo_id)
        reasons: list[str] = []

        if runtime is None:
            reasons.append("未完成运行检查。")
        else:
            if not runtime.checked:
                reasons.append("运行检查未完成。")
            if runtime.behavior_checked and not runtime.behavior_passed:
                reasons.extend(runtime.behavior_issues)
            elif not runtime.passed:
                reasons.extend(runtime.issues)

        reasons.extend(analysis.repo_issues.get(item.repo_id, []))

        normalized: list[str] = []
        for reason in reasons:
            text = _to_text(reason)
            if not text:
                continue
            if text in low_signal:
                continue
            if text not in normalized:
                normalized.append(text)

        if not normalized:
            normalized = ["未通过运行或行为验证。"]

        details.append(f"{item.repo_id}：{'；'.join(normalized[:2])}")

    return "2 分 repo 均未达到新增需求条件：" + " | ".join(details)


def validate_and_fix_decision(
    decision: AgentDecision,
    results: list[RepoResult],
    llm: ChatOpenAI,
    analysis: AnalysisBundle,
) -> AgentDecision:
    debug_repo = decision.debug_rewrite_repo.strip()
    new_repo = decision.new_requirement_rewrite_repo.strip()

    preferred_debug = _pick_debug_repo(results, analysis)
    selected_debug = _repo_by_id(debug_repo, results) if debug_repo else None
    if selected_debug is None or not debug_repo:
        debug_repo = preferred_debug
    else:
        if preferred_debug and not _has_runtime_bug(selected_debug.repo_id, analysis):
            debug_repo = preferred_debug

    eligible_score2 = _eligible_score2_repo_ids(results, analysis)
    selected_new = _repo_by_id(new_repo, results) if new_repo else None
    if (
        selected_new is None
        or selected_new.score != 2
        or selected_new.repo_id not in eligible_score2
    ):
        new_repo = eligible_score2[0] if eligible_score2 else ""

    debug_prompt = decision.debug_prompt.strip()
    selected_debug_obj = _repo_by_id(debug_repo, results) if debug_repo else None
    selected_verified = (
        _verified_bug_lines(selected_debug_obj, analysis) if selected_debug_obj else []
    )
    if not debug_prompt:
        debug_prompt = (
            _fallback_debug_prompt(results, debug_repo, analysis)
            if debug_repo
            else DEBUG_UNABLE_TEXT
        )
    if not debug_repo or not selected_verified:
        # 没有可验证缺陷时，禁止生成猜测型 debug prompt。
        debug_prompt = DEBUG_UNABLE_TEXT

    new_prompt = decision.new_requirement_prompt.strip()
    unable_note = decision.unable_to_add_requirement_note.strip()
    if new_repo:
        if not new_prompt:
            new_prompt = _fallback_new_requirement_prompt()
        if not unable_note:
            unable_note = "无。已选择通过行为验证的 2 分产物用于新增需求。"
    else:
        new_prompt = ""
        if not unable_note:
            score2_exists = any(r.score == 2 for r in results)
            if score2_exists:
                unable_note = _summarize_score2_failures(results, analysis)
            else:
                unable_note = "没有可用的 2 分 repo，无法新增需求。"

    if _has_repo_marker(debug_prompt) or _has_english(debug_prompt) or _has_colloquial(debug_prompt):
        debug_prompt = rewrite_prompt_to_formal_chinese(llm, debug_prompt, "调试提示词")
    if new_prompt and (_has_repo_marker(new_prompt) or _has_english(new_prompt) or _has_colloquial(new_prompt)):
        new_prompt = rewrite_prompt_to_formal_chinese(llm, new_prompt, "新增需求提示词")

    if _has_repo_marker(debug_prompt) or _has_english(debug_prompt) or _has_colloquial(debug_prompt):
        debug_prompt = (
            _fallback_debug_prompt(results, debug_repo, analysis)
            if debug_repo
            else DEBUG_UNABLE_TEXT
        )
    if new_prompt and (_has_repo_marker(new_prompt) or _has_english(new_prompt) or _has_colloquial(new_prompt)):
        new_prompt = _fallback_new_requirement_prompt()

    return AgentDecision(
        debug_prompt=debug_prompt,
        debug_rewrite_repo=debug_repo,
        new_requirement_prompt=new_prompt,
        new_requirement_rewrite_repo=new_repo,
        unable_to_add_requirement_note=unable_note,
    )


def _behavior_summary(runtime: RuntimeCheck | None) -> dict[str, Any]:
    if runtime is None:
        return {}
    evidence = runtime.behavior_evidence
    return {
        "behavior_checked": runtime.behavior_checked,
        "behavior_passed": runtime.behavior_passed,
        "behavior_issue_count": len(runtime.behavior_issues),
        "behavior_issues": runtime.behavior_issues,
        "remark_probe_checked": runtime.remark_probe_checked,
        "remark_related_issues": runtime.remark_related_issues,
        "behavior_evidence": evidence,
    }


def build_llm_input(
    parsed: ParsedInput,
    results: list[RepoResult],
    analysis: AnalysisBundle,
) -> str:
    cases: list[dict[str, Any]] = []
    for result in results:
        runtime = analysis.runtime_checks.get(result.repo_id)
        cases.append(
            {
                "repo_id": result.repo_id,
                "url": result.url,
                "score": result.score,
                "remark": result.remark,
                "fetch_ok": result.ok,
                "title": result.title,
                "text_preview": result.text_preview,
                "text_length": result.text_length,
                "error": result.error,
                "extra_issues": analysis.repo_issues.get(result.repo_id, []),
                "runtime_checked": runtime.checked if runtime else False,
                "runtime_passed": runtime.passed if runtime else False,
                "runtime_issues": runtime.issues if runtime else [],
                "behavior": _behavior_summary(runtime),
                "verified_bug_lines": _verified_bug_lines(result, analysis),
            }
        )

    payload = {
        "uid": parsed.uid,
        "base_prompt": parsed.base_prompt,
        "rules": {
            "selection_core": (
                "debug 只能基于已验证缺陷输出；"
                "先做备注定向验错，再做通用行为探测。"
            ),
            "debug_select_priority": (
                "有已验证缺陷+有remark > 有已验证缺陷且0/1分 > 其余有已验证缺陷。"
            ),
            "new_requirement_select": (
                "新增需求改写repo只能选择2分且运行与行为验证均通过的repo。"
            ),
            "prompt_style": "两个prompt必须中文、正式、具体、非口语化。",
            "prompt_forbidden": "两个prompt中不得出现repo编号、网址、英文。",
            "new_requirement_constraint": "新增需求必须是单一具体功能，且包含验收标准。",
            "debug_constraint": "debug_prompt不得描述未验证问题，只能使用verified_bug_lines。",
        },
        "global_extra_issues": analysis.global_issues,
        "cases": cases,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def parse_llm_json(content: str) -> dict[str, Any]:
    text = content.strip()
    fenced = re.search(r"```json\s*(\{[\s\S]*?\})\s*```", text)
    if fenced:
        text = fenced.group(1)
    parsed = json.loads(text)
    if not isinstance(parsed, dict):
        raise ValueError("LLM output must be a JSON object")
    return parsed


def decide_for_case(
    llm: ChatOpenAI,
    parsed: ParsedInput,
    results: list[RepoResult],
    analysis: AnalysisBundle,
) -> AgentDecision:
    system_prompt = (
        "你是用于生成质检提示词的Agent。"
        "必须输出JSON对象。"
        "请以行为验证结论为主要依据，评分和备注仅作辅助。"
        "debug_prompt只能基于已验证缺陷（verified_bug_lines、remark_related_issues、behavior_issues）生成。"
        "不得描述未验证或推测性问题。"
        "debug与新增需求各只选一个repo。"
        "新增需求改写repo只能从“2分且运行与行为验证均通过”的repo中选择；"
        "若无可选项必须留空并写明原因。"
        "两个prompt必须中文、正式、具体、非口语化，不得出现repo编号、网址和英文。"
        "新增需求prompt必须是一个具体功能，并包含验收标准。"
    )
    human_prompt = (
        "请根据以下输入返回JSON对象。"
        "只返回JSON，不要解释，不要代码块。\n"
        f"{build_llm_input(parsed, results, analysis)}"
    )
    response = llm.invoke(
        [SystemMessage(content=system_prompt), HumanMessage(content=human_prompt)]
    )
    raw = parse_llm_json(_to_text(response.content))

    normalized = {
        "debug_prompt": raw.get("debug_prompt", ""),
        "debug_rewrite_repo": raw.get("debug_rewrite_repo", raw.get("debug_repo_id", "")),
        "new_requirement_prompt": raw.get(
            "new_requirement_prompt",
            raw.get("new_requirement", {}).get("prompt", "")
            if isinstance(raw.get("new_requirement"), dict)
            else "",
        ),
        "new_requirement_rewrite_repo": raw.get(
            "new_requirement_rewrite_repo",
            raw.get(
                "new_requirement_repo_id",
                raw.get("new_requirement", {}).get("repo_id", "")
                if isinstance(raw.get("new_requirement"), dict)
                else "",
            ),
        ),
        "unable_to_add_requirement_note": raw.get(
            "unable_to_add_requirement_note",
            raw.get("unable_note", ""),
        ),
    }
    decision = AgentDecision.model_validate(normalized)
    return validate_and_fix_decision(decision, results, llm, analysis)


def _debug_unable_note(decision: AgentDecision) -> str:
    if not decision.debug_rewrite_repo.strip():
        return DEBUG_UNABLE_TEXT
    if decision.debug_prompt.strip() == DEBUG_UNABLE_TEXT:
        return DEBUG_UNABLE_TEXT
    return "无"


def _build_chat_model() -> ChatOpenAI:
    load_dotenv()
    base_url = _require_env("MODEL_BASE_URL")
    model_name = _require_env("MODEL_NAME")
    api_key = _require_env("MODEL_API_KEY")
    return ChatOpenAI(
        model=model_name,
        temperature=0.2,
        api_key=api_key,
        base_url=base_url,
    )


def process_excel(args: argparse.Namespace) -> None:
    input_path = Path(args.input_xlsx)
    if not input_path.exists():
        raise FileNotFoundError(f"Input Excel not found: {input_path}")

    output_path = Path(args.output_xlsx) if args.output_xlsx else input_path
    workbook = load_workbook(input_path)
    sheet = workbook[args.sheet_name] if args.sheet_name else workbook[workbook.sheetnames[0]]

    raw_map = _raw_header_map(sheet)
    input_cols = _resolve_input_columns(raw_map)
    output_cols = _ensure_output_columns(sheet, raw_map)

    llm = _build_chat_model()
    processed = 0
    debug_payloads: list[dict[str, Any]] = []

    for row_idx in range(args.start_row, sheet.max_row + 1):
        if args.max_rows > 0 and processed >= args.max_rows:
            break

        row_input = parse_row_input(sheet, row_idx, input_cols)
        if row_input is None:
            continue
        if args.uid and row_input.uid != args.uid:
            continue

        print(f"[Row {row_idx}] uid={row_input.uid}: fetch + behavior validation...")
        repo_results = fetch_repos(row_input)
        analysis = build_analysis(repo_results, row_input.base_prompt)

        print(f"[Row {row_idx}] uid={row_input.uid}: llm decision...")
        decision = decide_for_case(llm, row_input, repo_results, analysis)

        sheet.cell(row=row_idx, column=output_cols["debug_prompt"]).value = decision.debug_prompt
        sheet.cell(row=row_idx, column=output_cols["debug_repo"]).value = decision.debug_rewrite_repo
        sheet.cell(row=row_idx, column=output_cols["debug_unable_note"]).value = _debug_unable_note(decision)
        sheet.cell(row=row_idx, column=output_cols["new_prompt"]).value = decision.new_requirement_prompt
        sheet.cell(row=row_idx, column=output_cols["new_repo"]).value = decision.new_requirement_rewrite_repo
        sheet.cell(row=row_idx, column=output_cols["new_unable_note"]).value = (
            decision.unable_to_add_requirement_note
        )

        debug_payloads.append(
            {
                "row": row_idx,
                "uid": row_input.uid,
                "input": asdict(row_input),
                "results": [asdict(item) for item in repo_results],
                "analysis": {
                    "repo_issues": analysis.repo_issues,
                    "global_issues": analysis.global_issues,
                    "runtime_checks": {
                        key: asdict(value) for key, value in analysis.runtime_checks.items()
                    },
                },
                "decision": decision.model_dump(),
            }
        )

        processed += 1
        print(f"[Row {row_idx}] uid={row_input.uid}: done.")

    workbook.save(output_path)
    Path(args.fetch_json).write_text(
        json.dumps(debug_payloads, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Processed rows: {processed}")
    print(f"Excel written: {output_path}")
    print(f"Fetch/runtime JSON written: {args.fetch_json}")


def main() -> None:
    args = parse_args()
    process_excel(args)


if __name__ == "__main__":
    main()
